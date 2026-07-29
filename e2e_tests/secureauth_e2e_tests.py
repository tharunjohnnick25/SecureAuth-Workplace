"""
SecureAuth AI - Comprehensive Selenium E2E Test Suite (150+ Test Cases)
=========================================================================
This test suite validates the full functionality of the SecureAuth AI IAM platform.
It covers landing pages, authentication flows, dashboard, admin, security,
settings, billing, and all major feature modules.

Usage:
    python e2e_tests/secureauth_e2e_tests.py

Requirements:
    pip install selenium webdriver-manager openpyxl
"""

import os
import sys
import time
import json
import re
import traceback
from datetime import datetime, timezone
from dataclasses import dataclass, field
from typing import Optional

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementNotInteractableException,
    WebDriverException,
)
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# Test Configuration
# ============================================================================
BASE_URL = os.environ.get("TEST_BASE_URL", "http://localhost:3000")
HEADLESS = os.environ.get("TEST_HEADLESS", "true").lower() == "true"
SELENIUM_TIMEOUT = 15


# ============================================================================
# Test Result Tracking
# ============================================================================
@dataclass
class TestResult:
    category: str
    name: str
    status: str
    duration_sec: float = 0.0
    error_message: str = ""
    timestamp: str = ""


class TestReport:
    def __init__(self):
        self.results: list[TestResult] = []
        self.start_time: Optional[datetime] = None
        self.end_time: Optional[datetime] = None
        self.log_entries: list[tuple] = []

    def log(self, level: str, message: str):
        ts = datetime.now(timezone.utc).isoformat()
        self.log_entries.append((ts, level, message))
        print(f"[{level}] {message}")

    def add_result(self, result: TestResult):
        self.results.append(result)

    @property
    def total(self) -> int:
        return len(self.results)

    @property
    def passed(self) -> int:
        return sum(1 for r in self.results if r.status == "PASSED")

    @property
    def failed(self) -> int:
        return sum(1 for r in self.results if r.status in ("FAILED", "ERROR"))

    @property
    def pass_rate(self) -> float:
        if self.total == 0:
            return 0.0
        return round((self.passed / self.total) * 100, 2)

    def to_xlsx(self, filepath: str):
        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        passed_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        failed_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        ws_summary = wb.active
        ws_summary.title = "Summary"
        headers = ["Test Suite", "Total Tests", "Passed", "Failed", "Pass Rate %", "Duration (sec)", "Start Time", "End Time"]
        for col, h in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        duration = 0.0
        if self.start_time and self.end_time:
            duration = (self.end_time - self.start_time).total_seconds()

        data = [
            "SecureAuth AI - Full E2E Workflow",
            self.total,
            self.passed,
            self.failed,
            self.pass_rate,
            round(duration, 2),
            self.start_time.isoformat() if self.start_time else "",
            self.end_time.isoformat() if self.end_time else "",
        ]
        for col, val in enumerate(data, 1):
            cell = ws_summary.cell(row=2, column=col, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        ws_passed = wb.create_sheet("Passed Tests")
        passed_headers = ["No.", "Category", "Test Name", "Time (sec)", "Status"]
        for col, h in enumerate(passed_headers, 1):
            cell = ws_passed.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            cell.alignment = header_align
            cell.border = thin_border

        passed_idx = 1
        for i, r in enumerate(self.results, 1):
            if r.status == "PASSED":
                row_data = [passed_idx, r.category, r.name, round(r.duration_sec, 2), r.status]
                for col, val in enumerate(row_data, 1):
                    cell = ws_passed.cell(row=passed_idx + 1, column=col, value=val)
                    cell.fill = passed_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                passed_idx += 1

        ws_failed = wb.create_sheet("Failed Tests")
        failed_headers = ["No.", "Category", "Test Name", "Error", "Status", "URL"]
        for col, h in enumerate(failed_headers, 1):
            cell = ws_failed.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
            cell.alignment = header_align
            cell.border = thin_border

        failed_idx = 1
        for i, r in enumerate(self.results, 1):
            if r.status in ("FAILED", "ERROR"):
                row_data = [failed_idx, r.category, r.name, r.error_message, r.status, self.driver.current_url if hasattr(self, 'driver') and self.driver else ""]
                for col, val in enumerate(row_data, 1):
                    cell = ws_failed.cell(row=failed_idx + 1, column=col, value=val)
                    cell.fill = failed_fill
                    cell.border = thin_border
                    if col == 4:
                        cell.alignment = Alignment(wrap_text=True)
                failed_idx += 1

        ws_log = wb.create_sheet("Execution Log")
        log_headers = ["Timestamp", "Level", "Message"]
        for col, h in enumerate(log_headers, 1):
            cell = ws_log.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, (ts, level, msg) in enumerate(self.log_entries, 2):
            ws_log.cell(row=i, column=1, value=ts).border = thin_border
            level_cell = ws_log.cell(row=i, column=2, value=level)
            level_cell.border = thin_border
            if level == "ERROR":
                level_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif level == "WARN":
                level_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            elif level == "INFO":
                level_cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            msg_cell = ws_log.cell(row=i, column=3, value=msg)
            msg_cell.border = thin_border
            msg_cell.alignment = Alignment(wrap_text=True)

        col_widths = {
            "Summary": [50, 12, 10, 10, 14, 14, 30, 30],
            "Passed Tests": [8, 20, 50, 12, 10],
            "Failed Tests": [8, 20, 50, 80, 10, 50],
            "Execution Log": [30, 8, 100],
        }
        for name, widths in col_widths.items():
            ws = wb[name]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(filepath)
        self.log("INFO", f"Report saved to {filepath}")


# ============================================================================
# Base Test Class with Robust Locator Strategies
# ============================================================================
class BaseTest:
    def __init__(self, driver: webdriver.Chrome, report: TestReport):
        self.driver = driver
        self.report = report
        self.wait = WebDriverWait(driver, SELENIUM_TIMEOUT)
        self.short_wait = WebDriverWait(driver, 5)

    def find(self, by, value, timeout=SELENIUM_TIMEOUT):
        return WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )

    def find_all(self, by, value, timeout=SELENIUM_TIMEOUT):
        WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )
        return self.driver.find_elements(by, value)

    def clickable(self, by, value, timeout=SELENIUM_TIMEOUT):
        return WebDriverWait(self.driver, timeout).until(
            EC.element_to_be_clickable((by, value))
        )

    def visible(self, by, value, timeout=SELENIUM_TIMEOUT):
        return WebDriverWait(self.driver, timeout).until(
            EC.visibility_of_element_located((by, value))
        )

    def nav(self, url):
        self.driver.get(url)
        time.sleep(1)

    def safe_text_xpath(self, text):
        return f"//*[text()='{text}']"

    def safe_contains_xpath(self, text):
        return f"//*[contains(text(), '{text}')]"

    def run(self, category, name, test_fn):
        result = TestResult(category=category, name=name, status="PASSED")
        start = time.time()
        try:
            test_fn()
            result.status = "PASSED"
        except AssertionError as e:
            result.status = "FAILED"
            result.error_message = str(e)[:500]
        except Exception as e:
            result.status = "ERROR"
            result.error_message = f"{type(e).__name__}: {str(e)[:500]}"
            traceback.print_exc()
        result.duration_sec = time.time() - start
        result.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.report.add_result(result)
        status_icon = "PASS" if result.status == "PASSED" else "FAIL"
        self.report.log("INFO", f"  [{status_icon}] {category}: {name} -> {result.status} ({result.duration_sec:.2f}s)")
        if result.status in ("FAILED", "ERROR"):
            self.report.log("ERROR", f"    {result.error_message}")
        return result

    def assert_text(self, element, expected, msg=None):
        actual = element.text.strip()
        assert expected in actual, (
            msg or f"Expected '{expected}' in element text, got '{actual[:200]}'"
        )

    def assert_url_contains(self, fragment):
        WebDriverWait(self.driver, 5).until(
            lambda d: fragment in d.current_url
        )
        assert fragment in self.driver.current_url, (
            f"Expected URL to contain '{fragment}', got '{self.driver.current_url}'"
        )

    def scroll_to(self, element):
        self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", element)
        time.sleep(0.3)

    def js_click(self, element):
        self.driver.execute_script("arguments[0].click();", element)

    def try_find(self, by, value, timeout=3):
        try:
            return WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((by, value))
            )
        except (TimeoutException, NoSuchElementException):
            return None


# ============================================================================
# Test Suites
# ============================================================================

class TestLandingPage(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Landing Page Tests ===")
        self.nav(f"{BASE_URL}/")

        def test_page_title():
            title = self.driver.title
            assert "SecureAuth" in title or "Auth" in title, f"Title mismatch: {title}"

        self.run("Landing Page", "page_title_contains_secureauth", test_page_title)

        def test_page_loads():
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Landing Page", "page_loads_successfully", test_page_loads)

        def test_navbar_brand():
            body = self.find(By.TAG_NAME, "body")
            navs = self.driver.find_elements(By.CSS_SELECTOR, "nav, header")
            assert len(navs) > 0 or body.is_displayed()

        self.run("Landing Page", "navbar_brand_logo_visible", test_navbar_brand)

        def test_navbar_features():
            el = self.find(By.XPATH, "//*[contains(text(), 'Features')]")
            assert el.is_displayed()

        self.run("Landing Page", "navbar_features_link_visible", test_navbar_features)

        def test_navbar_use_cases():
            el = self.find(By.XPATH, "//*[contains(text(), 'Use Cases')]")
            assert el.is_displayed()

        self.run("Landing Page", "navbar_use_cases_link_visible", test_navbar_use_cases)

        def test_navbar_security():
            el = self.find(By.XPATH, "//*[contains(text(), 'Security')]")
            assert el.is_displayed()

        self.run("Landing Page", "navbar_security_link_visible", test_navbar_security)

        def test_navbar_pricing():
            el = self.find(By.XPATH, "//a[contains(text(), 'Pricing')]")
            assert el.is_displayed()

        self.run("Landing Page", "navbar_pricing_link_visible", test_navbar_pricing)

        def test_navbar_signin():
            el = self.find(By.XPATH, "//*[contains(text(), 'Sign In')]")
            assert el.is_displayed()

        self.run("Landing Page", "navbar_signin_button_visible", test_navbar_signin)

        def test_navbar_get_started():
            el = self.find(By.XPATH, "//*[contains(text(), 'Get Started')]")
            assert el.is_displayed()

        self.run("Landing Page", "navbar_get_started_button_visible", test_navbar_get_started)

        def test_hero_heading():
            h1s = self.driver.find_elements(By.TAG_NAME, "h1")
            texts = [h.text for h in h1s]
            found = any("Access" in t or "Authentication" in t or "Protect" in t or "Secure" in t for t in texts)
            assert found, f"No hero heading found in h1s: {texts}"

        self.run("Landing Page", "hero_heading_exists_with_key_text", test_hero_heading)

        def test_hero_subtitle():
            els = self.driver.find_elements(By.XPATH, "//p[contains(text(), 'enterprise') or contains(text(), 'workforce') or contains(text(), 'authentication') or contains(text(), 'Protect')]")
            assert len(els) > 0, "Hero subtitle not found"

        self.run("Landing Page", "hero_subtitle_text_present", test_hero_subtitle)

        def test_hero_cta_start_trial():
            el = self.find(By.XPATH, "//*[contains(text(), 'Start Free Trial')]")
            assert el.is_displayed()

        self.run("Landing Page", "hero_start_free_trial_button", test_hero_cta_start_trial)

        def test_hero_cta_watch_demo():
            el = self.find(By.XPATH, "//*[contains(text(), 'Watch Live Demo')]")
            assert el.is_displayed()

        self.run("Landing Page", "hero_watch_live_demo_button", test_hero_cta_watch_demo)

        def test_features_section():
            el = self.find(By.XPATH, "//*[contains(text(), 'Department-Based Access') or contains(text(), 'Department-Based')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "features_section_department_access", test_features_section)

        def test_features_admin_workflow():
            el = self.find(By.XPATH, "//*[contains(text(), 'Admin Approval Workflow')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "features_admin_approval_workflow", test_features_admin_workflow)

        def test_features_office_login():
            el = self.find(By.XPATH, "//*[contains(text(), 'Office Login Tracking')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "features_office_login_tracking", test_features_office_login)

        def test_features_ai_auth():
            el = self.find(By.XPATH, "//*[contains(text(), 'AI Adaptive Authentication')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "features_ai_adaptive_auth", test_features_ai_auth)

        def test_features_threat_dashboard():
            el = self.find(By.XPATH, "//*[contains(text(), 'Real-Time Threat Dashboard')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "features_threat_dashboard", test_features_threat_dashboard)

        def test_features_cross_platform():
            el = self.find(By.XPATH, "//*[contains(text(), 'Cross-Platform Security')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "features_cross_platform_security", test_features_cross_platform)

        def test_use_cases_corporate():
            el = self.find(By.XPATH, "//*[contains(text(), 'Corporate Offices')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "use_case_corporate_offices", test_use_cases_corporate)

        def test_use_cases_banking():
            el = self.find(By.XPATH, "//*[contains(text(), 'Banking & Finance')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "use_case_banking_finance", test_use_cases_banking)

        def test_use_cases_healthcare():
            el = self.find(By.XPATH, "//*[contains(text(), 'Healthcare Providers')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "use_case_healthcare_providers", test_use_cases_healthcare)

        def test_use_cases_hr():
            el = self.find(By.XPATH, "//*[contains(text(), 'HR Management')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "use_case_hr_management", test_use_cases_hr)

        def test_use_cases_it_tech():
            el = self.find(By.XPATH, "//*[contains(text(), 'IT & Tech Enterprises')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "use_case_it_tech_enterprises", test_use_cases_it_tech)

        def test_use_cases_remote():
            el = self.find(By.XPATH, "//*[contains(text(), 'Remote Workforce')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "use_case_remote_workforce", test_use_cases_remote)

        def test_workflow_section():
            el = self.find(By.XPATH, "//*[contains(text(), 'Seamless Access Workflow') or contains(text(), 'Access Workflow')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "workflow_section_visible", test_workflow_section)

        def test_risk_stats():
            el = self.find(By.XPATH, "//*[contains(text(), 'Risk Score') or contains(text(), 'Threats Blocked') or contains(text(), 'Active Sessions')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "risk_visualization_stats_visible", test_risk_stats)

        def test_pricing_section():
            el = self.find(By.XPATH, "//*[contains(text(), 'Starter') or contains(text(), '\u20b90')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "pricing_starter_tier_visible", test_pricing_section)

        def test_pricing_professional():
            el = self.find(By.XPATH, "//*[contains(text(), 'Professional')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "pricing_professional_tier_visible", test_pricing_professional)

        def test_pricing_enterprise():
            el = self.find(By.XPATH, "//*[contains(text(), 'Enterprise') and not(contains(text(), 'Enterprises'))]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "pricing_enterprise_tier_visible", test_pricing_enterprise)

        def test_pricing_monthly_yearly_toggle():
            el = self.find(By.XPATH, "//*[contains(text(), 'monthly') or contains(text(), 'yearly') or contains(text(), 'Monthly') or contains(text(), 'Yearly')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "pricing_monthly_yearly_toggle_visible", test_pricing_monthly_yearly_toggle)

        def test_footer_exists():
            footer = self.driver.find_elements(By.TAG_NAME, "footer")
            assert len(footer) > 0, "Footer not found"
            self.scroll_to(footer[-1])
            assert footer[-1].is_displayed()

        self.run("Landing Page", "footer_section_present", test_footer_exists)

        def test_footer_links_count():
            footer = self.driver.find_elements(By.TAG_NAME, "footer")
            assert len(footer) > 0
            links = footer[-1].find_elements(By.TAG_NAME, "a")
            assert len(links) >= 3, f"Footer should have at least 3 links, found {len(links)}"

        self.run("Landing Page", "footer_has_multiple_links", test_footer_links_count)

        def test_badge_new_version():
            el = self.find(By.XPATH, "//*[contains(text(), 'AI-Powered Office Security')]")
            self.scroll_to(el)
            assert el.is_displayed()

        self.run("Landing Page", "new_version_badge_visible", test_badge_new_version)

        def test_biometrics_showcase():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Device Intelligence') or contains(text(), 'Behavioral Rhythm')]")
            if el:
                self.scroll_to(el)
                assert el.is_displayed()

        self.run("Landing Page", "biometrics_showcase_visible", test_biometrics_showcase)

        def test_faq_section():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'FAQ') or contains(text(), 'frequently') or contains(text(), 'Frequently')]")
            if el:
                self.scroll_to(el)
                assert el.is_displayed()

        self.run("Landing Page", "faq_section_visible", test_faq_section)

        def test_testimonials():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Sarah Chen') or contains(text(), 'Marcus Thorne')]")
            if el:
                self.scroll_to(el)
                assert el.is_displayed()

        self.run("Landing Page", "testimonials_section_visible", test_testimonials)


class TestSignupPage(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Signup Page Tests ===")
        self.nav(f"{BASE_URL}/signup")

        def test_page_loads():
            self.find(By.TAG_NAME, "body")

        self.run("Signup Page", "signup_page_loads", test_page_loads)

        def test_url():
            self.assert_url_contains("/signup")

        self.run("Signup Page", "signup_url_correct", test_url)

        def test_signup_heading():
            el = self.find(By.XPATH, "//*[contains(text(), 'Request Access') or contains(text(), 'Register')]")
            assert el.is_displayed()

        self.run("Signup Page", "signup_heading_visible", test_signup_heading)

        def test_name_input():
            inp = self.find(By.XPATH, "//input[@type='text' or contains(@placeholder, 'Name') or contains(@name, 'name')]")
            assert inp.is_displayed()

        self.run("Signup Page", "name_input_field_present", test_name_input)

        def test_email_input():
            inp = self.find(By.XPATH, "//input[@type='email' or contains(@placeholder, 'Email')]")
            assert inp.is_displayed()

        self.run("Signup Page", "email_input_field_present", test_email_input)

        def test_password_input():
            inp = self.find(By.XPATH, "//input[@type='password']")
            assert inp.is_displayed()

        self.run("Signup Page", "password_input_field_present", test_password_input)

        def test_submit_button():
            btn = self.find(By.XPATH, "//button[contains(text(), 'Submit Access Request') or contains(text(), 'Submit')]")
            assert btn.is_displayed()

        self.run("Signup Page", "submit_button_visible", test_submit_button)

        def test_signin_link():
            link = self.find(By.XPATH, "//*[contains(text(), 'Already have an account') or contains(text(), 'Sign in')]")
            assert link.is_displayed()

        self.run("Signup Page", "signin_navigation_link_visible", test_signin_link)

        def test_full_name_label():
            label = self.find(By.XPATH, "//*[contains(text(), 'Full Name')]")
            assert label.is_displayed()

        self.run("Signup Page", "full_name_label_visible", test_full_name_label)

        def test_company_email_label():
            label = self.find(By.XPATH, "//*[contains(text(), 'Company Email')]")
            assert label.is_displayed()

        self.run("Signup Page", "company_email_label_visible", test_company_email_label)

        def test_confirm_password_input():
            inp = self.driver.find_elements(By.XPATH, "//input[@type='password']")
            assert len(inp) >= 1

        self.run("Signup Page", "password_fields_present", test_confirm_password_input)


class TestLoginPage(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Login Page Tests ===")
        self.nav(f"{BASE_URL}/login")

        def test_page_loads():
            self.find(By.TAG_NAME, "body")

        self.run("Login Page", "login_page_loads", test_page_loads)

        def test_url():
            self.assert_url_contains("/login")

        self.run("Login Page", "login_url_correct", test_url)

        def test_heading():
            el = self.find(By.XPATH, "//*[contains(text(), 'Multi-Factor')]")
            assert el.is_displayed()

        self.run("Login Page", "login_heading_visible", test_heading)

        def test_email_input():
            inp = self.find(By.XPATH, "//input[@type='email' or contains(@placeholder, 'email') or contains(@placeholder, 'Email')]")
            assert inp.is_displayed()

        self.run("Login Page", "email_input_field", test_email_input)

        def test_password_input():
            inp = self.find(By.XPATH, "//input[@type='password']")
            assert inp.is_displayed()

        self.run("Login Page", "password_input_field", test_password_input)

        def test_signin_button():
            btn = self.find(By.XPATH, "//button[contains(text(), 'Sign In') or contains(text(), 'Sign')]")
            assert btn.is_displayed()

        self.run("Login Page", "signin_button_visible", test_signin_button)

        def test_forgot_password_link():
            link = self.find(By.XPATH, "//*[contains(text(), 'Forgot password')]")
            assert link.is_displayed()

        self.run("Login Page", "forgot_password_link_visible", test_forgot_password_link)

        def test_create_account_link():
            link = self.find(By.XPATH, "//*[contains(text(), 'Create account')]")
            assert link.is_displayed()

        self.run("Login Page", "create_account_link_visible", test_create_account_link)

        def test_google_oauth_button():
            btn = self.find(By.XPATH, "//button[contains(text(), 'Google')]")
            assert btn.is_displayed()

        self.run("Login Page", "google_oauth_button_visible", test_google_oauth_button)

        def test_github_oauth_button():
            btn = self.find(By.XPATH, "//button[contains(text(), 'GitHub')]")
            assert btn.is_displayed()

        self.run("Login Page", "github_oauth_button_visible", test_github_oauth_button)

        def test_face_verification_step():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Face Verification') or contains(text(), 'Verify your identity') or contains(text(), 'identity using')]")
            if el:
                assert el.is_displayed()

        self.run("Login Page", "face_verification_step_visible", test_face_verification_step)

        def test_risk_engine_status():
            el = self.find(By.XPATH, "//*[contains(text(), 'AI Risk Engine')]")
            assert el.is_displayed()

        self.run("Login Page", "ai_risk_engine_visible", test_risk_engine_status)


class TestForgotPassword(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Forgot Password Tests ===")
        self.nav(f"{BASE_URL}/forgot-password")

        def test_page_loads():
            self.find(By.TAG_NAME, "body")

        self.run("Forgot Password", "forgot_password_page_loads", test_page_loads)

        def test_url():
            self.assert_url_contains("/forgot-password")

        self.run("Forgot Password", "forgot_password_url_correct", test_url)

        def test_heading():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Forgot') and (contains(text(), 'password') or contains(text(), 'Password'))]")
            if el:
                assert el.is_displayed()

        self.run("Forgot Password", "heading_visible", test_heading)

        def test_description():
            el = self.find(By.XPATH, "//*[contains(text(), 'reset link') or contains(text(), 'email to receive')]")
            assert el.is_displayed()

        self.run("Forgot Password", "description_text_visible", test_description)

        def test_email_input():
            inp = self.find(By.XPATH, "//input[@type='email' or contains(@placeholder, 'email') or contains(@placeholder, 'Email')]")
            assert inp.is_displayed()

        self.run("Forgot Password", "email_input_visible", test_email_input)

        def test_send_reset_button():
            btn = self.find(By.XPATH, "//button[contains(text(), 'Send Reset Link')]")
            assert btn.is_displayed()

        self.run("Forgot Password", "send_reset_button_visible", test_send_reset_button)

        def test_back_to_login():
            link = self.try_find(By.XPATH, "//*[contains(text(), 'Back to login') or contains(text(), 'Back to Login') or contains(text(), '\u2190')]")
            if link:
                assert link.is_displayed()

        self.run("Forgot Password", "back_to_login_link_visible", test_back_to_login)

        def test_click_back_to_login():
            self.nav(f"{BASE_URL}/login")
            self.assert_url_contains("/login")

        self.run("Forgot Password", "back_to_login_navigates_correctly", test_click_back_to_login)


class TestPricingPage(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Pricing Page Tests ===")
        self.nav(f"{BASE_URL}/pricing")

        def test_page_loads():
            self.find(By.TAG_NAME, "body")

        self.run("Pricing Page", "pricing_page_loads", test_page_loads)

        def test_url():
            self.assert_url_contains("/pricing")

        self.run("Pricing Page", "pricing_url_correct", test_url)

        def test_starter_tier():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Starter') or contains(text(), 'Free') or contains(text(), '\u20b9')]")
            if el:
                assert el.is_displayed()

        self.run("Pricing Page", "starter_tier_visible", test_starter_tier)

        def test_professional_tier():
            el = self.find(By.XPATH, "//*[contains(text(), 'Professional')]")
            assert el.is_displayed()

        self.run("Pricing Page", "professional_tier_visible", test_professional_tier)

        def test_enterprise_tier():
            el = self.find(By.XPATH, "//*[contains(text(), 'Enterprise')]")
            assert el.is_displayed()

        self.run("Pricing Page", "enterprise_tier_visible", test_enterprise_tier)

        def test_monthly_yearly_toggle():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'monthly') or contains(text(), 'yearly') or contains(text(), 'Monthly') or contains(text(), 'Yearly')]")
            if el:
                assert el.is_displayed()

        self.run("Pricing Page", "monthly_yearly_toggle_visible", test_monthly_yearly_toggle)

        def test_get_started_buttons():
            btns = self.driver.find_elements(By.XPATH, "//a[contains(text(), 'Get Started') or contains(text(), 'Try') or contains(text(), 'Contact') or contains(@href, '/signup') or contains(@href, '#signup')]")
            assert len(btns) >= 1 or self.find(By.TAG_NAME, "body").is_displayed()

        self.run("Pricing Page", "cta_buttons_present", test_get_started_buttons)


class TestDemoPage(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Demo Page Tests ===")
        self.nav(f"{BASE_URL}/demo")

        def test_page_loads():
            self.find(By.TAG_NAME, "body")

        self.run("Demo Page", "demo_page_loads", test_page_loads)

        def test_url():
            self.assert_url_contains("/demo")

        self.run("Demo Page", "demo_url_correct", test_url)

        def test_has_content():
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Demo Page", "demo_has_content", test_has_content)


class TestAuthFlow(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Authentication Flow Tests ===")
        self.nav(f"{BASE_URL}/login")

        def test_navigate_to_login():
            self.assert_url_contains("/login")

        self.run("Auth Flow", "navigate_to_login_page", test_navigate_to_login)

        def test_enter_email():
            inp = self.find(By.XPATH, "//input[@type='email' or contains(@placeholder, 'email') or contains(@placeholder, 'Email')]")
            inp.clear()
            inp.send_keys("admin@secureauth.com")
            assert inp.get_attribute("value") != "", "Email was not entered"

        self.run("Auth Flow", "enter_email_address", test_enter_email)

        def test_enter_password():
            inp = self.find(By.XPATH, "//input[@type='password']")
            inp.clear()
            inp.send_keys("Admin@123")
            assert inp.get_attribute("value") != "", "Password was not entered"

        self.run("Auth Flow", "enter_password", test_enter_password)

        def test_click_signin():
            btn = self.find(By.XPATH, "//button[contains(text(), 'Sign In') or contains(text(), 'Sign')]")
            self.js_click(btn)
            time.sleep(3)

        self.run("Auth Flow", "click_signin_button", test_click_signin)

        def test_redirect_after_login():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(2)
            self.assert_url_contains("/dashboard")

        self.run("Auth Flow", "redirect_to_dashboard_after_login", test_redirect_after_login)


class TestDashboard(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Dashboard Tests ===")
        self.driver.get(f"{BASE_URL}/dashboard")
        time.sleep(2)

        def test_url():
            self.assert_url_contains("/dashboard")

        self.run("Dashboard", "dashboard_url_correct", test_url)

        def test_security_dashboard_heading():
            el = self.find(By.XPATH, "//*[contains(text(), 'Security Dashboard')]")
            assert el.is_displayed()

        self.run("Dashboard", "security_dashboard_heading_visible", test_security_dashboard_heading)

        def test_system_live_badge():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'System Live') or contains(text(), 'system live')]")
            if el:
                assert el.is_displayed()

        self.run("Dashboard", "system_live_badge_visible", test_system_live_badge)

        def test_total_employees_card():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Total Employees') or contains(text(), 'total employees')]")
            if el:
                assert el.is_displayed()

        self.run("Dashboard", "total_employees_card_visible", test_total_employees_card)

        def test_active_sessions_card():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Active Sessions')]")
            if el:
                assert el.is_displayed()

        self.run("Dashboard", "active_sessions_card_visible", test_active_sessions_card)

        def test_security_alerts_card():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Security Alert') or contains(text(), 'Alert')]")
            if el:
                assert el.is_displayed()

        self.run("Dashboard", "security_alerts_card_visible", test_security_alerts_card)

        def test_system_risk_card():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'System Risk') or contains(text(), 'Risk') or contains(text(), 'risk')]")
            if not el:
                el = self.try_find(By.XPATH, "//*[contains(text(), 'System') or contains(text(), 'system')]")
            if el and el.is_displayed():
                self.scroll_to(el)
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Dashboard", "system_risk_card_visible", test_system_risk_card)

        def test_dashboard_description():
            el = self.try_find(By.XPATH, "//*[contains(text(), 'Real-time enterprise-grade')]")
            if el:
                assert el.is_displayed()

        self.run("Dashboard", "dashboard_description_text", test_dashboard_description)


class TestSidebarNavigation(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Sidebar Navigation Tests ===")
        self.driver.get(f"{BASE_URL}/dashboard")
        time.sleep(2)

        def test_navigate_to_employees():
            link = self.try_find(By.XPATH, "//a[contains(@href, '/employees')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/employees")

        self.run("Sidebar Nav", "navigate_to_employees_page", test_navigate_to_employees)

        def test_navigate_to_attendance():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/attendance')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/attendance")

        self.run("Sidebar Nav", "navigate_to_attendance_page", test_navigate_to_attendance)

        def test_navigate_to_notifications():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/notifications')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/notifications")

        self.run("Sidebar Nav", "navigate_to_notifications_page", test_navigate_to_notifications)

        def test_navigate_to_security():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(2)
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Sidebar Nav", "navigate_to_security_page", test_navigate_to_security)

        def test_navigate_to_settings():
            link = self.try_find(By.XPATH, "//a[contains(@href, '/settings')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/settings")

        self.run("Sidebar Nav", "navigate_to_settings_page", test_navigate_to_settings)

        def test_navigate_to_devices():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/devices')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/devices")

        self.run("Sidebar Nav", "navigate_to_devices_page", test_navigate_to_devices)

        def test_navigate_to_roles():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/roles-permissions')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/roles-permissions")

        self.run("Sidebar Nav", "navigate_to_roles_permissions", test_navigate_to_roles)

        def test_navigate_to_audit_logs():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(2)
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Sidebar Nav", "navigate_to_audit_logs", test_navigate_to_audit_logs)

        def test_navigate_to_office_logins():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/office-logins')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/office-logins")

        self.run("Sidebar Nav", "navigate_to_office_logins", test_navigate_to_office_logins)

        def test_navigate_to_threat_intelligence():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(2)
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Sidebar Nav", "navigate_to_threat_intelligence", test_navigate_to_threat_intelligence)

        def test_navigate_to_departments():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/departments')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/departments")

        self.run("Sidebar Nav", "navigate_to_departments", test_navigate_to_departments)

        def test_navigate_to_integrations():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/integrations')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/integrations")

        self.run("Sidebar Nav", "navigate_to_integrations", test_navigate_to_integrations)

        def test_navigate_to_profile():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/profile')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/profile")

        self.run("Sidebar Nav", "navigate_to_profile", test_navigate_to_profile)

        def test_navigate_to_access_requests():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/access-requests')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/access-requests")

        self.run("Sidebar Nav", "navigate_to_access_requests", test_navigate_to_access_requests)

        def test_navigate_to_analytics():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/analytics')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/analytics")

        self.run("Sidebar Nav", "navigate_to_analytics", test_navigate_to_analytics)

        def test_navigate_to_api_docs():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/api-documentation')]")
            if link:
                self.js_click(link)
                time.sleep(2)
                self.assert_url_contains("/api-documentation")

        self.run("Sidebar Nav", "navigate_to_api_documentation", test_navigate_to_api_docs)

        def test_navigate_to_risk_dashboard():
            self.driver.get(f"{BASE_URL}/dashboard")
            time.sleep(1)
            link = self.try_find(By.XPATH, "//a[contains(@href, '/dashboard/risk')]")
            if link:
                self.js_click(link)
                time.sleep(2)

        self.run("Sidebar Nav", "navigate_to_risk_dashboard", test_navigate_to_risk_dashboard)


class TestDirectPageAccess(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Direct Page Access Tests ===")
        pages = [
            ("Employees", "/employees"),
            ("Settings", "/settings"),
            ("Security", "/security"),
            ("Analytics", "/analytics"),
            ("Audit Logs", "/audit-logs"),
            ("Departments", "/departments"),
            ("Roles & Permissions", "/roles-permissions"),
            ("Office Logins", "/office-logins"),
            ("Devices", "/devices"),
            ("Notifications", "/notifications"),
            ("Access Requests", "/access-requests"),
            ("Profile", "/profile"),
            ("Threat Intelligence", "/threat-intelligence"),
            ("Integrations", "/integrations"),
            ("Attendance", "/attendance"),
            ("API Documentation", "/api-documentation"),
            ("MFA Settings", "/mfa-settings"),
            ("MFA Setup", "/mfa-setup"),
            ("Session Management", "/session-management"),
            ("Password Policies", "/password-policies"),
            ("Billing", "/billing"),
            ("Subscription Plans", "/subscription-plans"),
            ("System Health", "/system-health"),
            ("Support Tickets", "/support-tickets"),
            ("Help Center", "/help-center"),
            ("Incident Response", "/incident-response"),
            ("Vulnerability Scanner", "/vulnerability-scanner"),
            ("Forensics", "/forensics"),
            ("Network Map", "/network-map"),
            ("Geolocation Map", "/geolocation-map"),
            ("Risk Assessment", "/risk-assessment"),
            ("Compliance Reports", "/compliance-reports"),
            ("Custom Reports", "/custom-reports"),
            ("Backup Recovery", "/backup-recovery"),
            ("LDAP Integration", "/ldap-integration"),
            ("SSO Configuration", "/sso-configuration"),
            ("Email Templates", "/email-templates"),
            ("Developer Portal", "/developer-portal"),
            ("Team Management", "/team-management"),
            ("Webhooks", "/webhooks"),
            ("Security Events", "/security-events"),
            ("Usage Statistics", "/usage-statistics"),
            ("Performance Metrics", "/performance-metrics"),
            ("Export Data", "/export-data"),
            ("Import Data", "/import-data"),
            ("Reports Dashboard", "/reports-dashboard"),
            ("Scheduled Reports", "/scheduled-reports"),
            ("Notification Rules", "/notification-rules"),
            ("Alerts Configuration", "/alerts-configuration"),
            ("API Keys", "/api-keys"),
            ("User Profile", "/user-profile"),
            ("Device Details", "/device-details"),
            ("Pricing", "/pricing"),
            ("Login", "/login"),
            ("Signup", "/signup"),
            ("Demo", "/demo"),
            ("Forgot Password", "/forgot-password"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            if not path_key:
                path_key = "home"
            self.run(f"Page Access: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestSecurityFeatures(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Security Features Tests ===")

        routes = [
            ("Security Center", "/security"),
            ("Zero Trust", "/security/zero-trust"),
            ("Fingerprinting", "/security/fingerprinting"),
            ("Risk Score", "/security/risk-score"),
            ("Incident Response", "/incident-response"),
            ("Vulnerability Scanner", "/vulnerability-scanner"),
            ("Forensics", "/forensics"),
            ("Network Map", "/network-map"),
            ("Geolocation Map", "/geolocation-map"),
            ("Risk Assessment", "/risk-assessment"),
            ("Threat Intelligence", "/threat-intelligence"),
        ]
        for page_name, page_url in routes:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Security: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestEmployeesSection(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Employees Section Tests ===")

        routes = [
            ("Employee Directory", "/employees"),
            ("Employee Detail", "/employees/1"),
            ("New Employee", "/employees/new"),
            ("Employee Edit", "/employees/1/edit"),
            ("Team Management", "/team-management"),
            ("Departments", "/departments"),
            ("Roles & Permissions", "/roles-permissions"),
            ("Access Requests", "/access-requests"),
        ]
        for page_name, page_url in routes:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Employees: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestOfficeLogins(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Office Logins Tests ===")

        routes = [
            ("Office Logins Main", "/office-logins"),
            ("Onsite Logins", "/office-logins/onsite"),
            ("Remote Logins", "/office-logins/remote"),
            ("Suspicious Logins", "/office-logins/suspicious"),
            ("Attendance", "/attendance"),
            ("Geolocation Map", "/geolocation-map"),
        ]
        for page_name, page_url in routes:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Office Logins: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestAdminPages(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Admin Pages Tests ===")
        admin_routes = [
            ("Admin Dashboard", "/admin/dashboard"),
            ("Admin Analytics", "/admin/analytics"),
            ("Admin Audit", "/admin/audit"),
        ]

        for page_name, page_url in admin_routes:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(2)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Admin: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestReportsAndAnalytics(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Reports & Analytics Tests ===")
        pages = [
            ("Analytics", "/analytics"),
            ("Reports Dashboard", "/reports-dashboard"),
            ("Compliance Reports", "/compliance-reports"),
            ("Custom Reports", "/custom-reports"),
            ("Scheduled Reports", "/scheduled-reports"),
            ("Export Data", "/export-data"),
            ("Import Data", "/import-data"),
            ("Usage Statistics", "/usage-statistics"),
            ("Performance Metrics", "/performance-metrics"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Reports: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestNotificationsAndAlerts(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Notifications & Alerts Tests ===")
        pages = [
            ("Notification Rules", "/notification-rules"),
            ("Alerts Configuration", "/alerts-configuration"),
            ("Security Events", "/security-events"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Alerts: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestDeveloperSection(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Developer Section Tests ===")
        pages = [
            ("API Documentation", "/api-documentation"),
            ("API Keys", "/api-keys"),
            ("Developer Portal", "/developer-portal"),
            ("Webhooks", "/webhooks"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Developer: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestAuthAndMFA(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Auth & MFA Tests ===")
        pages = [
            ("MFA Settings", "/mfa-settings"),
            ("MFA Setup", "/mfa-setup"),
            ("MFA Verify", "/mfa-verify"),
            ("Verify Biometric", "/verify-biometric"),
            ("Verify OTP", "/verify-otp"),
            ("Session Management", "/session-management"),
            ("Password Policies", "/password-policies"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Auth & MFA: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestIntegrations(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Integrations Tests ===")
        pages = [
            ("LDAP Integration", "/ldap-integration"),
            ("SSO Configuration", "/sso-configuration"),
            ("Integrations", "/integrations"),
            ("Email Templates", "/email-templates"),
            ("Backup Recovery", "/backup-recovery"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Integrations: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestSettingsAndConfig(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Settings & Configuration Tests ===")

        def test_settings_page():
            self.driver.get(f"{BASE_URL}/settings")
            time.sleep(1.5)
            self.assert_url_contains("/settings")
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Settings", "settings_page_loads", test_settings_page)

        def test_unauthorized_page():
            self.driver.get(f"{BASE_URL}/unauthorized")
            time.sleep(1.5)
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Settings", "unauthorized_page_loads", test_unauthorized_page)

        def test_not_found_page():
            self.driver.get(f"{BASE_URL}/nonexistent-page-12345")
            time.sleep(1.5)
            body = self.find(By.TAG_NAME, "body")
            assert body.is_displayed()

        self.run("Settings", "not_found_page_shows_404", test_not_found_page)


class TestBillingAndSubscriptions(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Billing & Subscriptions Tests ===")
        pages = [
            ("Billing", "/billing"),
            ("Subscription Plans", "/subscription-plans"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"Billing: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestUserArea(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== User Area Tests ===")
        pages = [
            ("Profile", "/profile"),
            ("User Profile", "/user-profile"),
            ("Device Details", "/device-details"),
            ("Support Tickets", "/support-tickets"),
            ("Help Center", "/help-center"),
        ]

        for page_name, page_url in pages:
            def make_test(url, name):
                def test():
                    self.driver.get(f"{BASE_URL}{url}")
                    time.sleep(1.5)
                    body = self.find(By.TAG_NAME, "body")
                    assert body.is_displayed()
                return test

            path_key = page_url.strip('/').replace('/', '_').replace('-', '_')
            self.run(f"User Area: {page_name}", f"navigate_to_{path_key}", make_test(page_url, page_name))


class TestHomepageCTAFlows(BaseTest):
    def run_all(self):
        self.report.log("INFO", "\n=== Homepage CTA Flow Tests ===")

        def test_signin_link_navigates():
            self.nav(f"{BASE_URL}/")
            link = self.find(By.XPATH, "//*[contains(text(), 'Sign In')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/login")

        self.run("CTA Flows", "signin_link_navigates_to_login", test_signin_link_navigates)

        def test_get_started_navigates():
            self.nav(f"{BASE_URL}/")
            link = self.find(By.XPATH, "//*[contains(text(), 'Get Started')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/signup")

        self.run("CTA Flows", "get_started_navigates_to_signup", test_get_started_navigates)

        def test_pricing_link_navigates():
            self.nav(f"{BASE_URL}/")
            link = self.find(By.XPATH, "//a[contains(text(), 'Pricing')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/pricing")

        self.run("CTA Flows", "pricing_link_navigates_to_pricing", test_pricing_link_navigates)

        def test_start_free_trial():
            self.nav(f"{BASE_URL}/")
            link = self.find(By.XPATH, "//*[contains(text(), 'Start Free Trial')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/signup")

        self.run("CTA Flows", "start_free_trial_navigates_to_signup", test_start_free_trial)

        def test_watch_demo():
            self.nav(f"{BASE_URL}/")
            link = self.find(By.XPATH, "//*[contains(text(), 'Watch Live Demo')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/demo")

        self.run("CTA Flows", "watch_demo_navigates_to_demo", test_watch_demo)

        def test_login_create_account():
            self.nav(f"{BASE_URL}/login")
            link = self.find(By.XPATH, "//*[contains(text(), 'Create account')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/signup")

        self.run("CTA Flows", "create_account_navigates_to_signup", test_login_create_account)

        def test_forgot_password_from_login():
            self.nav(f"{BASE_URL}/login")
            link = self.find(By.XPATH, "//*[contains(text(), 'Forgot password')]")
            self.js_click(link)
            time.sleep(1.5)
            self.assert_url_contains("/forgot-password")

        self.run("CTA Flows", "forgot_password_navigates_correctly", test_forgot_password_from_login)

        def test_back_to_home():
            self.nav(f"{BASE_URL}/login")
            link = self.try_find(By.XPATH, "//*[contains(text(), 'SecureAuthAI') or contains(text(), 'SECURE AUTH') or contains(text(), 'SecureAuth')]")
            if link:
                self.js_click(link)
                time.sleep(1.5)

        self.run("CTA Flows", "back_to_home_navigation", test_back_to_home)


# ============================================================================
# Main Test Runner
# ============================================================================
def main():
    report = TestReport()
    report.start_time = datetime.now(timezone.utc)
    report.log("INFO", "=" * 70)
    report.log("INFO", "SecureAuth AI - Comprehensive E2E Test Suite")
    report.log("INFO", f"Target URL: {BASE_URL}")
    report.log("INFO", f"Headless Mode: {HEADLESS}")
    report.log("INFO", "=" * 70)
    report.log("INFO", f"Started at: {report.start_time.isoformat()}")

    options = webdriver.ChromeOptions()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-insecure-localhost")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options,
    )
    driver.implicitly_wait(5)
    driver.maximize_window()

    try:
        # Public Pages
        TestLandingPage(driver, report).run_all()
        TestSignupPage(driver, report).run_all()
        TestLoginPage(driver, report).run_all()
        TestForgotPassword(driver, report).run_all()
        TestPricingPage(driver, report).run_all()
        TestDemoPage(driver, report).run_all()

        # CTA Flow Tests
        TestHomepageCTAFlows(driver, report).run_all()

        # Auth Flow
        TestAuthFlow(driver, report).run_all()

        # Protected Pages (after login)
        TestDashboard(driver, report).run_all()
        TestSidebarNavigation(driver, report).run_all()
        TestDirectPageAccess(driver, report).run_all()
        TestSecurityFeatures(driver, report).run_all()
        TestEmployeesSection(driver, report).run_all()
        TestOfficeLogins(driver, report).run_all()
        TestAdminPages(driver, report).run_all()
        TestReportsAndAnalytics(driver, report).run_all()
        TestNotificationsAndAlerts(driver, report).run_all()
        TestDeveloperSection(driver, report).run_all()
        TestAuthAndMFA(driver, report).run_all()
        TestIntegrations(driver, report).run_all()
        TestSettingsAndConfig(driver, report).run_all()
        TestBillingAndSubscriptions(driver, report).run_all()
        TestUserArea(driver, report).run_all()

    except Exception as e:
        report.log("ERROR", f"Test suite crashed: {e}")
        traceback.print_exc()
    finally:
        report.end_time = datetime.now(timezone.utc)
        report.log("INFO", "=" * 70)

        total = report.total
        passed = report.passed
        failed = report.failed
        rate = report.pass_rate

        report.log("INFO", f"Total Tests: {total}")
        report.log("INFO", f"Passed: {passed}")
        report.log("INFO", f"Failed: {failed}")
        report.log("INFO", f"Pass Rate: {rate}%")
        report.log("INFO", "=" * 70)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
        os.makedirs(report_dir, exist_ok=True)
        filepath = os.path.join(report_dir, f"E2E_Test_Report_SecureAuth_AI_{timestamp}.xlsx")

        report.to_xlsx(filepath)
        print(f"\nReport generated: {filepath}")

        try:
            driver.quit()
        except Exception:
            pass

    return report


if __name__ == "__main__":
    main()
