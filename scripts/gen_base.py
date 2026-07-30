import openpyxl, os
H = ['Test Case ID','Test Suite/Feature','Test Description','Preconditions','Test Steps','Test Data/Input','Expected Result','Actual Result','Status','Priority','Severity']
P = lambda n: str(n).zfill(3)

def save(fn, sn, rows):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = sn
    for ci,h in enumerate(H,1): ws.cell(1,ci,h)
    for ri,r in enumerate(rows,2):
        for ci,v in enumerate(r,1): ws.cell(ri,ci,v)
    ws.auto_filter.ref = f'A1:K{len(rows)+1}'
    wb.save(os.path.join(os.getcwd(), fn))
    wb.save(os.path.join('D:\\Downloads', fn))
    print(f'SAVED: {fn} ({len(rows)} cases)')

print("Generating 6 comprehensive test reports with 300 cases each...")
print("="*60)
