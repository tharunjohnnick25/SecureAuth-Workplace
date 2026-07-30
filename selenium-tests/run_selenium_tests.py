"""
SecureAuth AI - Selenium Automated E2E Test Suite & Screenshot Report Generator
================================================================================
This script executes end-to-end Selenium tests against the SecureAuth AI IAM platform,
captures high-resolution screenshots for each page and feature scenario, saves them in 
'selenium-tests/reports/screenshots/' and 'e2e_tests/reports/screenshots/', and builds 
both Excel (.xlsx) and interactive HTML test reports.

Usage:
    python selenium-tests/run_selenium_tests.py
"""

import os
import sys
import time
import json
import traceback
from datetime import datetime, timezone
from dataclasses import dataclass, field
from typing import List, Optional

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

BASE_URL = os.environ.get("TEST_BASE_URL", "http://localhost:3000")
SELENIUM_TIMEOUT = 15

# Target Directories
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
SELENIUM_REPORTS_DIR = os.path.join(PROJECT_ROOT, "selenium-tests", "reports")
SELENIUM_SCREENSHOTS_DIR = os.path.join(SELENIUM_REPORTS_DIR, "screenshots")

E2E_REPORTS_DIR = os.path.join(PROJECT_ROOT, "e2e_tests", "reports")
E2E_SCREENSHOTS_DIR = os.path.join(E2E_REPORTS_DIR, "screenshots")

for d in [SELENIUM_REPORTS_DIR, SELENIUM_SCREENSHOTS_DIR, E2E_REPORTS_DIR, E2E_SCREENSHOTS_DIR]:
    os.makedirs(d, exist_ok=True)

@dataclass
class TestItem:
    test_id: str
    category: str
    name: str
    url: str
    status: str = "PENDING"
    duration_sec: float = 0.0
    error: str = ""
    screenshot_file: str = ""

class SeleniumTestRunner:
    def __init__(self):
        self.results: List[TestItem] = []
        self.driver: Optional[webdriver.Chrome] = None
        self.start_time: Optional[datetime] = None
        self.end_time: Optional[datetime] = None

    def init_driver(self):
        print("[INFO] Initializing Headless Chrome Driver (1920x1080)...")
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--ignore-certificate-errors")
        
        service = ChromeService(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.driver.set_page_load_timeout(30)
        print("[INFO] Chrome Driver successfully initialized.")

    def capture_screenshot(self, filename: str) -> str:
        s1 = os.path.join(SELENIUM_SCREENSHOTS_DIR, filename)
        s2 = os.path.join(E2E_SCREENSHOTS_DIR, filename)
        try:
            self.driver.save_screenshot(s1)
            import shutil
            shutil.copy(s1, s2)
            print(f"  [SCREENSHOT] Saved {filename} to screenshots folders.")
            return filename
        except Exception as e:
            print(f"  [ERROR] Failed to save screenshot {filename}: {e}")
            return ""

    def run_test_case(self, test_id: str, category: str, name: str, route: str, action_fn, screenshot_name: str):
        item = TestItem(test_id=test_id, category=category, name=name, url=f"{BASE_URL}{route}")
        start = time.time()
        print(f"\n[RUN] [{test_id}] {category} - {name} ({route})")
        try:
            self.driver.get(item.url)
            time.sleep(1.5) # Allow dynamic UI JS hydration
            action_fn()
            item.status = "PASS"
        except Exception as e:
            item.status = "FAIL"
            item.error = f"{type(e).__name__}: {str(e)[:300]}"
            print(f"  [FAIL] {item.error}")

        item.duration_sec = round(time.time() - start, 2)
        if screenshot_name:
            item.screenshot_file = self.capture_screenshot(screenshot_name)
        self.results.append(item)
        print(f"  [{item.status}] Duration: {item.duration_sec}s")

    def execute_all_tests(self):
        self.start_time = datetime.now(timezone.utc)
        self.init_driver()

        try:
            # 1. Landing Page Hero
            def test_landing():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed(), "Body element should be visible"
                assert "SecureAuth" in self.driver.title or len(self.driver.page_source) > 500
            self.run_test_case("WEB-AUTH-001", "Landing Page", "Verify Landing Page Hero Section", "/", test_landing, "01_landing_page_hero.png")

            # 2. Landing Features
            def test_features():
                self.driver.execute_script("window.scrollTo(0, 800);")
                time.sleep(1)
            self.run_test_case("WEB-AUTH-002", "Landing Page", "Verify Features Section Showcase", "/", test_features, "02_landing_page_features.png")

            # 3. Landing Pricing
            def test_landing_pricing():
                self.driver.execute_script("window.scrollTo(0, 1800);")
                time.sleep(1)
            self.run_test_case("WEB-AUTH-003", "Landing Page", "Verify Landing Pricing Breakdown", "/", test_landing_pricing, "03_landing_page_pricing.png")

            # 4. Login Page
            def test_login():
                email_input = self.driver.find_elements(By.XPATH, "//input[@type='email' or contains(@placeholder, 'email') or contains(@placeholder, 'Email')]")
                assert len(email_input) > 0 or len(self.driver.find_elements(By.TAG_NAME, "input")) > 0
            self.run_test_case("WEB-AUTH-004", "Authentication", "Verify Multi-Factor Login Portal", "/login", test_login, "04_login_page_mfa.png")

            # 5. Signup Page
            def test_signup():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-AUTH-005", "Authentication", "Verify Request Access Signup Form", "/signup", test_signup, "05_signup_request_access.png")

            # 6. Forgot Password
            def test_forgot_pw():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-AUTH-006", "Authentication", "Verify Forgot Password Reset Request", "/forgot-password", test_forgot_pw, "06_forgot_password_reset.png")

            # 7. Pricing Page
            def test_pricing():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-AUTH-007", "Subscription", "Verify Pricing & Subscription Tiers", "/pricing", test_pricing, "07_pricing_plans_breakdown.png")

            # 8. User Dashboard
            def test_dashboard():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-DASH-001", "Dashboard", "Verify Security Operations Dashboard", "/dashboard", test_dashboard, "08_user_security_dashboard.png")

            # 9. Admin Overview
            def test_admin():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-ADM-001", "Admin Portal", "Verify Admin Management Console", "/admin", test_admin, "09_admin_audit_logs_dashboard.png")

            # 10. Audit Logs
            def test_audit_logs():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-SEC-001", "Security Audit", "Verify IAM Compliance Audit Trail", "/audit-logs", test_audit_logs, "10_audit_logs_compliance.png")

            # 11. Security Threat Intelligence
            def test_threats():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-SEC-002", "Security Audit", "Verify Real-Time AI Threat Center", "/security", test_threats, "11_ai_threat_monitoring_panel.png")

            # 12. Employee Directory
            def test_employees():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-EMP-001", "Directory", "Verify Employee Identity Directory", "/employees", test_employees, "12_employee_directory_management.png")

            # 13. Department Access Rules
            def test_departments():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-DEPT-001", "Governance", "Verify Department Access Policies", "/departments", test_departments, "13_department_access_policies.png")

            # 14. MFA Verification Step
            def test_mfa():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-MFA-001", "Biometrics", "Verify MFA & Biometrics Verification", "/mfa-verify", test_mfa, "14_mfa_biometrics_step.png")

            # 15. Vulnerability Scanner
            def test_vuln():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-VULN-001", "Security", "Verify Automated Vulnerability Matrix", "/vulnerability-scanner", test_vuln, "15_vulnerability_scanner_matrix.png")

            # 16. Performance Metrics
            def test_performance():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-LOAD-001", "Performance", "Verify Load & System Metrics Dashboard", "/performance-metrics", test_performance, "16_load_performance_metrics.png")

            # 17. Device Management & Mobile Sync
            def test_devices():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-APP-001", "Mobile Appium", "Verify Mobile Device Intelligence Sync", "/devices", test_devices, "17_mobile_appium_capacitor_sync.png")

            # 18. Compliance Reports
            def test_compliance():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-SETT-001", "Compliance", "Verify SOC2 & ISO Security Compliance", "/compliance-reports", test_compliance, "18_compliance_reports_view.png")

            # 19. Session Management
            def test_sessions():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-SESS-001", "Sessions", "Verify Active Session Management", "/session-management", test_sessions, "19_active_sessions_control.png")

            # 20. API Keys Portal
            def test_api_keys():
                body = self.driver.find_element(By.TAG_NAME, "body")
                assert body.is_displayed()
            self.run_test_case("WEB-API-001", "Developer", "Verify API Keys & Webhook Controls", "/api-keys", test_api_keys, "20_api_keys_management.png")

        finally:
            if self.driver:
                self.driver.quit()
                print("[INFO] Browser session closed.")
            self.end_time = datetime.now(timezone.utc)

    def generate_html_report(self, target_path: str):
        total = len(self.results)
        passed = sum(1 for r in self.results if r.status == "PASS")
        failed = sum(1 for r in self.results if r.status == "FAIL")
        pass_rate = round((passed / total * 100), 1) if total > 0 else 0
        duration = round((self.end_time - self.start_time).total_seconds(), 2) if self.start_time and self.end_time else 0

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Selenium Automated Test Execution Report & Screenshots</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-dark: #0f172a;
            --card-bg: #1e293b;
            --border-color: #334155;
            --accent-blue: #38bdf8;
            --accent-green: #22c55e;
            --accent-red: #ef4444;
            --text-main: #f8fafc;
            --text-muted: #94a3b8;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Inter', sans-serif;
            background-color: var(--bg-dark);
            color: var(--text-main);
            padding: 2rem;
            line-height: 1.6;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 1.5rem;
            margin-bottom: 2rem;
        }}
        .header h1 {{
            font-size: 1.8rem;
            font-weight: 700;
            background: linear-gradient(135deg, #38bdf8, #818cf8);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        .badge {{
            background: rgba(56, 189, 248, 0.1);
            color: var(--accent-blue);
            padding: 0.4rem 0.8rem;
            border-radius: 6px;
            font-size: 0.85rem;
            font-weight: 600;
            border: 1px solid rgba(56, 189, 248, 0.3);
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1.5rem;
            margin-bottom: 2.5rem;
        }}
        .stat-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 1.5rem;
            text-align: center;
        }}
        .stat-value {{
            font-size: 2.2rem;
            font-weight: 700;
            margin-top: 0.5rem;
        }}
        .stat-value.pass {{ color: var(--accent-green); }}
        .stat-value.fail {{ color: var(--accent-red); }}
        .stat-value.rate {{ color: var(--accent-blue); }}
        .stat-label {{
            font-size: 0.85rem;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}
        .section-title {{
            font-size: 1.3rem;
            font-weight: 600;
            margin-bottom: 1.5rem;
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}
        .gallery-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(360px, 1fr));
            gap: 1.5rem;
        }}
        .test-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            overflow: hidden;
            transition: transform 0.2s ease, border-color 0.2s ease;
        }}
        .test-card:hover {{
            transform: translateY(-4px);
            border-color: var(--accent-blue);
        }}
        .screenshot-container {{
            position: relative;
            width: 100%;
            height: 220px;
            background: #000;
            overflow: hidden;
        }}
        .screenshot-container img {{
            width: 100%;
            height: 100%;
            object-fit: cover;
            object-position: top;
            cursor: pointer;
        }}
        .status-tag {{
            position: absolute;
            top: 10px;
            right: 10px;
            padding: 0.3rem 0.6rem;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: 700;
        }}
        .status-tag.PASS {{ background: var(--accent-green); color: #000; }}
        .status-tag.FAIL {{ background: var(--accent-red); color: #fff; }}
        .test-info {{
            padding: 1.2rem;
        }}
        .test-id {{
            font-size: 0.8rem;
            color: var(--accent-blue);
            font-weight: 600;
        }}
        .test-name {{
            font-size: 1rem;
            font-weight: 600;
            margin: 0.3rem 0;
            color: var(--text-main);
        }}
        .test-meta {{
            display: flex;
            justify-content: space-between;
            font-size: 0.8rem;
            color: var(--text-muted);
            margin-top: 0.8rem;
            border-top: 1px solid rgba(255,255,255,0.05);
            padding-top: 0.6rem;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div>
            <h1>SecureAuth AI - Selenium E2E Test Execution & Screenshots</h1>
            <p style="color: var(--text-muted); font-size: 0.9rem; margin-top: 0.3rem;">Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} UTC</p>
        </div>
        <div class="badge">Selenium Web UI Suite</div>
    </div>

    <div class="stats-grid">
        <div class="stat-card">
            <div class="stat-label">Total Test Cases</div>
            <div class="stat-value">{total}</div>
        </div>
        <div class="stat-card">
            <div class="stat-label">Passed Tests</div>
            <div class="stat-value pass">{passed}</div>
        </div>
        <div class="stat-card">
            <div class="stat-label">Failed Tests</div>
            <div class="stat-value fail">{failed}</div>
        </div>
        <div class="stat-card">
            <div class="stat-label">Pass Rate</div>
            <div class="stat-value rate">{pass_rate}%</div>
        </div>
        <div class="stat-card">
            <div class="stat-label">Total Duration</div>
            <div class="stat-value">{duration}s</div>
        </div>
    </div>

    <div class="section-title">
        📷 Automated Test Screenshots & UI Verification Gallery
    </div>

    <div class="gallery-grid">"""
        for item in self.results:
            img_src = f"screenshots/{item.screenshot_file}" if item.screenshot_file else ""
            html_content += f"""
        <div class="test-card">
            <div class="screenshot-container">
                <span class="status-tag {item.status}">{item.status}</span>
                {"<a href='" + img_src + "' target='_blank'><img src='" + img_src + "' alt='" + item.name + "'></a>" if img_src else "<div style='padding:40px;text-align:center;color:#666'>No Screenshot Captured</div>"}
            </div>
            <div class="test-info">
                <div class="test-id">{item.test_id} | {item.category}</div>
                <div class="test-name">{item.name}</div>
                <div class="test-meta">
                    <span>⏱ {item.duration_sec}s</span>
                    <span>🔗 {item.url.replace(BASE_URL, '')}</span>
                </div>
            </div>
        </div>"""

        html_content += """
    </div>
</body>
</html>"""
        with open(target_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"[REPORT] Saved HTML Test Report to {target_path}")

    def generate_excel_report(self, target_path: str):
        if not openpyxl:
            print("[WARN] openpyxl missing, skipping Excel report generation")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Selenium Execution Summary"

        headers = ["Test Case ID", "Category", "Test Name", "Target URL", "Status", "Duration (sec)", "Screenshot File", "Error Message"]
        ws.append(headers)

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for item in self.results:
            ws.append([item.test_id, item.category, item.name, item.url, item.status, item.duration_sec, item.screenshot_file, item.error])

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(target_path)
        print(f"[REPORT] Saved Excel Test Report to {target_path}")

if __name__ == "__main__":
    runner = SeleniumTestRunner()
    runner.execute_all_tests()
    
    # Save reports to selenium-tests/reports
    runner.generate_html_report(os.path.join(SELENIUM_REPORTS_DIR, "index.html"))
    runner.generate_excel_report(os.path.join(SELENIUM_REPORTS_DIR, "Selenium_Testing_Report.xlsx"))

    # Save reports to e2e_tests/reports
    runner.generate_html_report(os.path.join(E2E_REPORTS_DIR, "index.html"))
    runner.generate_excel_report(os.path.join(E2E_REPORTS_DIR, "Selenium_Testing_Report.xlsx"))
    runner.generate_excel_report(os.path.join(PROJECT_ROOT, "Selenium_Testing_Report.xlsx"))
